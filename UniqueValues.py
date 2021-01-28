import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook

fileinput= input('What is the name of the file you want an ECM exception Report for?: ')
df = pd.read_excel(filepath+fileinput, skiprows= 1)
df.rename(columns ={'Buyer (User Pref)' : 'Buyer', 'Oper Unit': 'OperUnit', 'GL Unit': 'GL_Unit'}, inplace=True)
group = df.groupby('Dept')
df1 = group.apply(lambda x: x['Buyer'].unique())
chart = df[(df.OperUnit.isnull()) | (df.GL_Unit.isnull()) | (df.Fund.isnull()) | (df.Program.isnull()) | (df.Project.isnull())]
df1.to_excel(filepath)
df2 = pd.read_excel(filepath)
df2.rename(columns = {0 : 'Buyers'}, inplace= True)
buyers = df2.columns[1]
df2.to_excel(filepath, index=False)
df3 = pd.read_excel(filepath, index=False)
df3.iloc[(df3.Buyers.str.len()).argsort()].reset_index(drop=True)

df3['Buyer'] = df3['Buyers'].str.len()
df3.sort_values('Buyer', ascending=False, inplace=True)
df3.to_excel(filepath, index=False)
file_loc = r'filepath'
ss = openpyxl.load_workbook(file_loc)
ss_sheet1 = ss['Sheet1']
ss_sheet1.title = "Dept Buyers"
ss.create_sheet('BUYER01')
ss.create_sheet('Chartstring Errors')
ss.save(file_loc)

workbook1 = openpyxl.load_workbook(filepath)
writer = pd.ExcelWriter(filepath, engine='openpyxl')

new_file = df[(df.Buyer.isnull()) | (df.Buyer =='KENNEE03') | (df.Buyer == 'BUYER01')]
df3.to_excel(writer, sheet_name='DEPT BUYERS',index=False,startrow=0,startcol=0)
new_file.to_excel(writer, sheet_name='Incorrect Buyers OR Blanks',index=False,startrow=0,startcol=0)
chart.to_excel(writer, sheet_name='Chartstring Errors', index=False, startrow=0,startcol=0)
writer.save()
print('File Name--- ' + fileinput)
date =fileinput[:10]

os.rename(r'filepath', r'renamed filepath')

print('Exception Report Completed Successfully')
